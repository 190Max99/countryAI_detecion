"""
批量验证每个场景的得分准确率，并生成 Excel 报告。

功能：
1. 扫描 data/raw 下每个农户文件夹；
2. 自动读取每户 CSV，计算人工实际扣分/得分；
3. 自动读取五类图片，调用训练好的 ResNet18 / ResNet18+CBAM 模型预测扣分；
4. 统计每个场景的得分准确率、1分内准确率、2分内准确率、MAE、RMSE、平均分差；
5. 生成 Excel 文件，包含每户得分明细、每张图片明细、场景准确率汇总、单图扣分准确率、标签概率明细。

运行方式：
    python -m src.eval_scene_score_accuracy_excel --root data/raw --output outputs/scene_score_accuracy_report.xlsx

注意：
    需要安装 openpyxl：pip install openpyxl
"""

import argparse
import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
from PIL import Image

import torch
import torch.nn as nn
from torchvision import models, transforms


# =========================
# 1. 场景配置
# =========================

SCENE_CONFIGS = {
    "室内": {
        "aliases": ["室内"],
        "model_path": "models/indoor_resnet18.pth",
        "fallback_model_paths": [],
        "base_score": 10,
        "label_cols": [f"label_{i}" for i in range(10)],
        "label_names": [
            "A0_室内格局杂乱无章",
            "A1_室内家具摆放杂乱无章",
            "A2_室内生活用品杂乱无章",
            "A3_鸡鸭进入屋内共居",
            "A4_室内地面存在鸡鸭粪污",
            "A5_鸡跳在室内桌子上",
            "A6_室内地面垃圾乱丢现象严重",
            "A7_室内桌面沙发表面乱堆乱摆",
            "A8_室内墙面污迹不堪",
            "A9_室内其他脏乱情况",
        ],
        "deducts": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
        "thresholds": [0.55, 0.55, 0.55, 0.45, 0.45, 0.45, 0.50, 0.50, 0.50, 0.65],
    },
    "庭院": {
        "aliases": ["庭院"],
        "model_path": "models/courtyard_resnet18_cbam.pth",
        "fallback_model_paths": ["models/courtyard_resnet18.pth"],
        "base_score": 30,
        "label_cols": [f"label_{i}" for i in range(12)],
        "label_names": [
            "B0_庭院内生产工具杂乱无章",
            "B1_庭院内交通用具杂乱无章",
            "B2_庭院内其他情况杂乱无章",
            "B3_鸡鸭进入庭院乱跑",
            "B4_庭院内鸡粪鸭粪满地",
            "B5_庭院内地面垃圾乱丢现象严重",
            "B6_庭院内柴草堆码无序",
            "B7_庭院内柴草堆码不整齐",
            "B8_庭院内房屋立面乱挂乱画",
            "B9_庭院内搭建棚库破败物品堆码杂乱不堪",
            "B10_庭院污水横流",
            "B11_庭院内其他情况",
        ],
        "deducts": [3, 3, 2, 2, 2, 5, 1, 2, 2, 5, 2, 1],
        "thresholds": [0.55, 0.55, 0.60, 0.45, 0.45, 0.50, 0.55, 0.55, 0.55, 0.55, 0.50, 0.65],
    },
    "厕所": {
        "aliases": ["厕所", "厕屋"],
        "model_path": "models/toilet_resnet18.pth",
        "fallback_model_paths": [],
        "base_score": None,
        "label_cols": [f"label_{i}" for i in range(2)],
        "label_names": [
            "C0_厕屋脏乱",
            "C1_厕屋功能配备不齐全",
        ],
        "deducts": [2, 3],
        "thresholds": [0.50, 0.50],
    },
    "化粪池": {
        "aliases": ["化粪池"],
        "model_path": "models/septic_resnet18.pth",
        "fallback_model_paths": [],
        "base_score": None,
        "label_cols": [f"label_{i}" for i in range(3)],
        "label_names": [
            "C2_化粪池盖板挪开未关闭取粪口未关闭",
            "C3_化粪池粪污溢流",
            "C4_厕所周围其他情况",
        ],
        "deducts": [2, 2, 1],
        "thresholds": [0.50, 0.50, 0.65],
    },
    "房前屋后": {
        "aliases": ["房前屋后", "房屋前后", "房前屋后及2侧", "房前屋后及两侧", "屋后", "两侧", "2侧"],
        "model_path": "models/outside_resnet18.pth",
        "fallback_model_paths": [],
        "base_score": 10,
        "label_cols": [f"label_{i}" for i in range(5)],
        "label_names": [
            "D0_房屋旁柴草堆码乱堆不整齐",
            "D1_房屋周身存在污水横流现象",
            "D2_房屋周身瓜果棚架破败不堪",
            "D3_房屋周身鸡鸭棚圈破败不堪脏臭",
            "D4_房屋周身其他情况",
        ],
        "deducts": [3, 2, 2, 2, 1],
        "thresholds": [0.55, 0.50, 0.50, 0.50, 0.65],
    },
}

SCORE_SCENES = ["室内", "庭院", "厕所及化粪池", "房前屋后", "总分"]
IMAGE_SCENES = ["室内", "庭院", "厕所", "化粪池", "房前屋后"]
IMAGE_EXTS = [".jpg", ".jpeg", ".png", ".bmp", ".webp"]
MODEL_CACHE = {}


# =========================
# 2. CBAM 模型结构
# =========================

class ChannelAttention(nn.Module):
    def __init__(self, in_channels, reduction=16):
        super().__init__()
        hidden_channels = max(in_channels // reduction, 1)
        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        self.max_pool = nn.AdaptiveMaxPool2d(1)
        self.mlp = nn.Sequential(
            nn.Conv2d(in_channels, hidden_channels, kernel_size=1, bias=False),
            nn.ReLU(inplace=True),
            nn.Conv2d(hidden_channels, in_channels, kernel_size=1, bias=False),
        )
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg_out = self.mlp(self.avg_pool(x))
        max_out = self.mlp(self.max_pool(x))
        attention = self.sigmoid(avg_out + max_out)
        return x * attention


class SpatialAttention(nn.Module):
    def __init__(self, kernel_size=7):
        super().__init__()
        padding = kernel_size // 2
        self.conv = nn.Conv2d(2, 1, kernel_size=kernel_size, padding=padding, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg_out = torch.mean(x, dim=1, keepdim=True)
        max_out, _ = torch.max(x, dim=1, keepdim=True)
        attention_input = torch.cat([avg_out, max_out], dim=1)
        attention = self.sigmoid(self.conv(attention_input))
        return x * attention


class CBAM(nn.Module):
    def __init__(self, in_channels, reduction=16, kernel_size=7):
        super().__init__()
        self.channel_attention = ChannelAttention(in_channels, reduction)
        self.spatial_attention = SpatialAttention(kernel_size)

    def forward(self, x):
        x = self.channel_attention(x)
        x = self.spatial_attention(x)
        return x


class ResNet18CBAM(nn.Module):
    def __init__(self, num_labels):
        super().__init__()
        base_model = models.resnet18(weights=None)
        self.conv1 = base_model.conv1
        self.bn1 = base_model.bn1
        self.relu = base_model.relu
        self.maxpool = base_model.maxpool
        self.layer1 = base_model.layer1
        self.layer2 = base_model.layer2
        self.layer3 = base_model.layer3
        self.layer4 = base_model.layer4
        self.cbam = CBAM(in_channels=512)
        self.avgpool = base_model.avgpool
        self.fc = nn.Linear(512, num_labels)

    def forward(self, x):
        x = self.conv1(x)
        x = self.bn1(x)
        x = self.relu(x)
        x = self.maxpool(x)
        x = self.layer1(x)
        x = self.layer2(x)
        x = self.layer3(x)
        x = self.layer4(x)
        x = self.cbam(x)
        x = self.avgpool(x)
        x = torch.flatten(x, 1)
        x = self.fc(x)
        return x


# =========================
# 3. 通用工具函数
# =========================

def normalize_text(value):
    if pd.isna(value):
        return ""
    text = str(value)
    text = text.replace("\ufeff", "")
    text = text.replace("\t", "")
    text = text.replace(" ", "")
    text = text.replace("\u3000", "")
    return text.strip()


def sort_house_id(value):
    text = str(value).strip()
    try:
        return int(text)
    except Exception:
        return text


def read_csv_safely(csv_path: Path):
    encodings = ["utf-8-sig", "gbk", "gb18030", "utf-8"]
    seps = [None, "\t", ","]
    best_df = None
    best_col_count = 0

    for enc in encodings:
        for sep in seps:
            try:
                if sep is None:
                    df = pd.read_csv(csv_path, encoding=enc, sep=None, engine="python")
                else:
                    df = pd.read_csv(csv_path, encoding=enc, sep=sep)
                df.columns = [normalize_text(c) for c in df.columns]
                if len(df.columns) > best_col_count:
                    best_df = df
                    best_col_count = len(df.columns)
                if len(df.columns) >= 3:
                    return df
            except Exception:
                pass

    if best_df is not None:
        return best_df
    raise RuntimeError(f"无法读取 CSV 文件: {csv_path}")


def find_label_csv(folder: Path):
    preferred = folder / f"{folder.name}.csv"
    if preferred.exists():
        return preferred

    csv_files = sorted(folder.glob("*.csv"))
    bad_keywords = ["result", "detail", "score", "预测", "结果", "simple", "ui", "summary", "report"]
    valid_files = []

    for file in csv_files:
        name = file.name.lower()
        if any(k.lower() in name for k in bad_keywords):
            continue
        valid_files.append(file)

    if valid_files:
        return valid_files[0]
    if csv_files:
        return csv_files[0]
    return None


def prepare_label_df(label_csv: Path):
    df = read_csv_safely(label_csv)
    if "scene" in df.columns:
        scene_col = "scene"
    elif "序号" in df.columns:
        scene_col = "序号"
    elif "类别" in df.columns:
        scene_col = "类别"
    else:
        scene_col = df.columns[0]
    df["scene_norm"] = df[scene_col].apply(normalize_text)
    return df


def find_true_row(label_df: pd.DataFrame, scene_key: str):
    aliases = SCENE_CONFIGS[scene_key]["aliases"]
    rows = label_df[label_df["scene_norm"].isin(aliases)]
    if len(rows) > 0:
        return rows.iloc[0]

    for _, row in label_df.iterrows():
        scene_text = str(row["scene_norm"])
        for alias in aliases:
            if alias in scene_text:
                return row
    return None


def get_label_value(row, col):
    if row is None:
        return 0
    if col not in row.index:
        return 0
    value = row[col]
    if pd.isna(value):
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def calc_deduct_from_row(row, scene_key):
    cfg = SCENE_CONFIGS[scene_key]
    total_deduct = 0
    labels = []
    for col, deduct in zip(cfg["label_cols"], cfg["deducts"]):
        flag = get_label_value(row, col)
        labels.append(flag)
        if flag == 1:
            total_deduct += int(deduct)
    return total_deduct, np.array(labels, dtype=int)


def calc_score(base_score, deduct):
    if pd.isna(deduct):
        return np.nan
    return max(0, float(base_score) - float(deduct))


def find_image_in_folder(folder: Path, scene_key: str):
    aliases = SCENE_CONFIGS[scene_key]["aliases"]
    candidates = []

    for file in folder.iterdir():
        if not file.is_file():
            continue
        if file.suffix.lower() not in IMAGE_EXTS:
            continue
        filename = normalize_text(file.name)
        for alias in aliases:
            if alias in filename:
                candidates.append(file)
                break

    if not candidates:
        return None
    candidates = sorted(candidates, key=lambda p: p.name)
    return candidates[0]


# =========================
# 4. 模型加载与预测
# =========================

def resolve_model_path(scene_key):
    cfg = SCENE_CONFIGS[scene_key]
    candidates = [cfg["model_path"]] + cfg.get("fallback_model_paths", [])
    for path_text in candidates:
        path = Path(path_text)
        if path.exists():
            return path
    raise FileNotFoundError(f"{scene_key} 模型不存在，已尝试: {candidates}")


def build_plain_resnet18(num_labels):
    model = models.resnet18(weights=None)
    in_features = model.fc.in_features
    model.fc = nn.Linear(in_features, num_labels)
    return model


def build_model(num_labels, use_cbam=False):
    if use_cbam:
        return ResNet18CBAM(num_labels=num_labels)
    return build_plain_resnet18(num_labels=num_labels)


def get_transform():
    return transforms.Compose([
        transforms.Resize((224, 224)),
        transforms.ToTensor(),
        transforms.Normalize(
            mean=[0.485, 0.456, 0.406],
            std=[0.229, 0.224, 0.225]
        )
    ])


def load_checkpoint(model_path, device):
    try:
        return torch.load(model_path, map_location=device, weights_only=False)
    except TypeError:
        return torch.load(model_path, map_location=device)


def detect_cbam_model(checkpoint):
    if checkpoint.get("model_type", "") == "resnet18_cbam":
        return True
    state_dict = checkpoint.get("model_state_dict", {})
    return any(key.startswith("cbam.") for key in state_dict.keys())


def load_model_for_scene(scene_key, device):
    cache_key = (scene_key, str(device))
    if cache_key in MODEL_CACHE:
        return MODEL_CACHE[cache_key]

    cfg = SCENE_CONFIGS[scene_key]
    model_path = resolve_model_path(scene_key)
    checkpoint = load_checkpoint(model_path, device)

    deducts = checkpoint.get("deducts", cfg["deducts"])
    thresholds = np.array(checkpoint.get("thresholds", cfg["thresholds"]), dtype=float)
    label_names = checkpoint.get("label_names", cfg["label_names"])
    use_cbam = detect_cbam_model(checkpoint)

    model = build_model(num_labels=len(label_names), use_cbam=use_cbam)
    model.load_state_dict(checkpoint["model_state_dict"])
    model.to(device)
    model.eval()

    pack = {
        "model": model,
        "deducts": deducts,
        "thresholds": thresholds,
        "label_names": label_names,
        "use_cbam": use_cbam,
        "model_path": str(model_path),
    }
    MODEL_CACHE[cache_key] = pack
    return pack


def predict_scene(scene_key, image_path, device):
    model_pack = load_model_for_scene(scene_key, device)
    model = model_pack["model"]
    deducts = model_pack["deducts"]
    thresholds = model_pack["thresholds"]
    label_names = model_pack["label_names"]

    image = Image.open(image_path).convert("RGB")
    image_tensor = get_transform()(image).unsqueeze(0).to(device)

    with torch.no_grad():
        logits = model(image_tensor)
        probs = torch.sigmoid(logits)[0].detach().cpu().numpy()

    pred_labels = (probs >= thresholds).astype(int)
    pred_deduct = sum(int(deduct) for flag, deduct in zip(pred_labels, deducts) if int(flag) == 1)

    return {
        "pred_deduct": pred_deduct,
        "pred_labels": pred_labels,
        "probs": probs,
        "thresholds": thresholds,
        "deducts": deducts,
        "label_names": label_names,
        "model_path": model_pack["model_path"],
        "use_cbam": model_pack["use_cbam"],
    }


# =========================
# 5. 单户验证与批量汇总
# =========================

def eval_one_house(folder: Path, device):
    label_csv = find_label_csv(folder)
    if label_csv is None:
        return None, [], [], f"未找到CSV: {folder}"

    label_df = prepare_label_df(label_csv)

    image_rows = []
    label_rows = []
    scene_tmp = {}

    for scene_key in IMAGE_SCENES:
        cfg = SCENE_CONFIGS[scene_key]
        true_row = find_true_row(label_df, scene_key)
        image_path = find_image_in_folder(folder, scene_key)

        true_deduct, true_labels = calc_deduct_from_row(true_row, scene_key) if true_row is not None else (np.nan, None)

        if image_path is None:
            pred_info = None
            pred_deduct = np.nan
            pred_labels = None
            probs = None
            note = "未找到图片"
        else:
            pred_info = predict_scene(scene_key, image_path, device)
            pred_deduct = pred_info["pred_deduct"]
            pred_labels = pred_info["pred_labels"]
            probs = pred_info["probs"]
            note = ""

        deduct_error = np.nan if pd.isna(true_deduct) or pd.isna(pred_deduct) else float(pred_deduct) - float(true_deduct)
        deduct_abs_error = np.nan if pd.isna(deduct_error) else abs(deduct_error)
        deduct_exact = np.nan if pd.isna(deduct_error) else int(deduct_error == 0)

        # 单独的厕所/化粪池不直接计分，只记录扣分；其他场景记录得分
        base_score = cfg["base_score"]
        if base_score is None:
            true_score = np.nan
            pred_score = np.nan
            score_error = np.nan
            score_abs_error = np.nan
            score_exact = np.nan
        else:
            true_score = calc_score(base_score, true_deduct)
            pred_score = calc_score(base_score, pred_deduct)
            score_error = np.nan if pd.isna(true_score) or pd.isna(pred_score) else pred_score - true_score
            score_abs_error = np.nan if pd.isna(score_error) else abs(score_error)
            score_exact = np.nan if pd.isna(score_error) else int(score_error == 0)

        image_rows.append({
            "house_id": folder.name,
            "场景": scene_key,
            "图片路径": str(image_path) if image_path is not None else "",
            "标签CSV": str(label_csv),
            "模型路径": pred_info["model_path"] if pred_info is not None else "",
            "是否CBAM": pred_info["use_cbam"] if pred_info is not None else "",
            "实际扣分": true_deduct,
            "预测扣分": pred_deduct,
            "扣分误差": deduct_error,
            "扣分绝对误差": deduct_abs_error,
            "扣分完全准确": deduct_exact,
            "满分": base_score if base_score is not None else "不单独计分",
            "实际得分": true_score,
            "预测得分": pred_score,
            "得分误差": score_error,
            "得分绝对误差": score_abs_error,
            "得分完全准确": score_exact,
            "备注": note,
        })

        if pred_info is not None:
            for i, label_name in enumerate(pred_info["label_names"]):
                true_flag = int(true_labels[i]) if true_labels is not None and i < len(true_labels) else np.nan
                pred_flag = int(pred_labels[i]) if pred_labels is not None and i < len(pred_labels) else np.nan
                prob = float(probs[i]) if probs is not None and i < len(probs) else np.nan
                threshold = float(pred_info["thresholds"][i]) if i < len(pred_info["thresholds"]) else np.nan
                deduct = int(pred_info["deducts"][i]) if i < len(pred_info["deducts"]) else np.nan
                label_rows.append({
                    "house_id": folder.name,
                    "场景": scene_key,
                    "标签序号": i,
                    "标签名称": label_name,
                    "扣分值": deduct,
                    "实际标签": true_flag,
                    "预测标签": pred_flag,
                    "预测概率": prob,
                    "阈值": threshold,
                    "标签是否正确": np.nan if pd.isna(true_flag) or pd.isna(pred_flag) else int(true_flag == pred_flag),
                    "图片路径": str(image_path),
                })

        scene_tmp[scene_key] = {
            "true_deduct": true_deduct,
            "pred_deduct": pred_deduct,
            "true_score": true_score,
            "pred_score": pred_score,
        }

    # 合并厕所及化粪池得分
    toilet_true = scene_tmp.get("厕所", {}).get("true_deduct", np.nan)
    septic_true = scene_tmp.get("化粪池", {}).get("true_deduct", np.nan)
    toilet_pred = scene_tmp.get("厕所", {}).get("pred_deduct", np.nan)
    septic_pred = scene_tmp.get("化粪池", {}).get("pred_deduct", np.nan)

    if pd.isna(toilet_true) or pd.isna(septic_true):
        toilet_septic_true_deduct = np.nan
        toilet_septic_true_score = np.nan
    else:
        toilet_septic_true_deduct = float(toilet_true) + float(septic_true)
        toilet_septic_true_score = calc_score(10, toilet_septic_true_deduct)

    if pd.isna(toilet_pred) or pd.isna(septic_pred):
        toilet_septic_pred_deduct = np.nan
        toilet_septic_pred_score = np.nan
    else:
        toilet_septic_pred_deduct = float(toilet_pred) + float(septic_pred)
        toilet_septic_pred_score = calc_score(10, toilet_septic_pred_deduct)

    score_rows = []

    def append_score_row(scene_name, base_score, true_deduct, pred_deduct, true_score, pred_score):
        score_error = np.nan if pd.isna(true_score) or pd.isna(pred_score) else pred_score - true_score
        score_abs_error = np.nan if pd.isna(score_error) else abs(score_error)
        score_rows.append({
            "house_id": folder.name,
            "项目": scene_name,
            "满分": base_score,
            "实际扣分": true_deduct,
            "预测扣分": pred_deduct,
            "实际得分": true_score,
            "预测得分": pred_score,
            "分差": score_error,
            "绝对误差": score_abs_error,
            "得分完全准确": np.nan if pd.isna(score_error) else int(score_error == 0),
            "1分内准确": np.nan if pd.isna(score_abs_error) else int(score_abs_error <= 1),
            "2分内准确": np.nan if pd.isna(score_abs_error) else int(score_abs_error <= 2),
            "3分内准确": np.nan if pd.isna(score_abs_error) else int(score_abs_error <= 3),
            "5分内准确": np.nan if pd.isna(score_abs_error) else int(score_abs_error <= 5),
        })

    append_score_row(
        "室内", 10,
        scene_tmp["室内"]["true_deduct"], scene_tmp["室内"]["pred_deduct"],
        scene_tmp["室内"]["true_score"], scene_tmp["室内"]["pred_score"],
    )
    append_score_row(
        "庭院", 30,
        scene_tmp["庭院"]["true_deduct"], scene_tmp["庭院"]["pred_deduct"],
        scene_tmp["庭院"]["true_score"], scene_tmp["庭院"]["pred_score"],
    )
    append_score_row(
        "厕所及化粪池", 10,
        toilet_septic_true_deduct, toilet_septic_pred_deduct,
        toilet_septic_true_score, toilet_septic_pred_score,
    )
    append_score_row(
        "房前屋后", 10,
        scene_tmp["房前屋后"]["true_deduct"], scene_tmp["房前屋后"]["pred_deduct"],
        scene_tmp["房前屋后"]["true_score"], scene_tmp["房前屋后"]["pred_score"],
    )

    score_df = pd.DataFrame(score_rows)
    if score_df["实际得分"].isna().any() or score_df["预测得分"].isna().any():
        total_true = np.nan
        total_pred = np.nan
    else:
        total_true = float(score_df["实际得分"].sum())
        total_pred = float(score_df["预测得分"].sum())

    append_score_row(
        "总分", 60,
        np.nan if pd.isna(total_true) else 60 - total_true,
        np.nan if pd.isna(total_pred) else 60 - total_pred,
        total_true, total_pred,
    )

    return score_rows, image_rows, label_rows, ""


def collect_all_results(root_dir: Path, device):
    score_rows_all = []
    image_rows_all = []
    label_rows_all = []
    skipped_rows = []

    folders = [p for p in root_dir.iterdir() if p.is_dir()]
    folders = sorted(folders, key=lambda p: sort_house_id(p.name))

    for idx, folder in enumerate(folders, start=1):
        print(f"[{idx}/{len(folders)}] 正在验证: {folder}")
        try:
            score_rows, image_rows, label_rows, error = eval_one_house(folder, device)
            if error:
                skipped_rows.append({"house_id": folder.name, "文件夹": str(folder), "原因": error})
                print("  跳过：", error)
                continue
            score_rows_all.extend(score_rows)
            image_rows_all.extend(image_rows)
            label_rows_all.extend(label_rows)
        except Exception as e:
            skipped_rows.append({"house_id": folder.name, "文件夹": str(folder), "原因": str(e)})
            print("  出错，已跳过：", e)

    return (
        pd.DataFrame(score_rows_all),
        pd.DataFrame(image_rows_all),
        pd.DataFrame(label_rows_all),
        pd.DataFrame(skipped_rows),
    )


# =========================
# 6. 统计指标
# =========================

def rmse(series):
    values = pd.to_numeric(series, errors="coerce").dropna().values
    if len(values) == 0:
        return np.nan
    return float(np.sqrt(np.mean(values ** 2)))


def make_score_summary(score_detail_df: pd.DataFrame):
    rows = []
    for scene in SCORE_SCENES:
        df = score_detail_df[score_detail_df["项目"] == scene].copy()
        df = df.dropna(subset=["分差", "绝对误差"])
        n = len(df)
        if n == 0:
            continue
        rows.append({
            "项目": scene,
            "样本数": n,
            "得分完全准确数": int((df["得分完全准确"] == 1).sum()),
            "得分准确率": float((df["得分完全准确"] == 1).mean()),
            "1分内准确率": float((df["1分内准确"] == 1).mean()),
            "2分内准确率": float((df["2分内准确"] == 1).mean()),
            "3分内准确率": float((df["3分内准确"] == 1).mean()),
            "5分内准确率": float((df["5分内准确"] == 1).mean()),
            "MAE平均绝对误差": float(df["绝对误差"].mean()),
            "RMSE均方根误差": rmse(df["分差"]),
            "平均分差": float(df["分差"].mean()),
            "AI给分偏高数": int((df["分差"] > 0).sum()),
            "AI给分偏低数": int((df["分差"] < 0).sum()),
            "最高实际得分": float(df["实际得分"].max()),
            "最低实际得分": float(df["实际得分"].min()),
            "平均实际得分": float(df["实际得分"].mean()),
            "平均预测得分": float(df["预测得分"].mean()),
        })
    return pd.DataFrame(rows)


def make_deduct_summary(image_detail_df: pd.DataFrame):
    rows = []
    for scene in IMAGE_SCENES:
        df = image_detail_df[image_detail_df["场景"] == scene].copy()
        df = df.dropna(subset=["扣分误差", "扣分绝对误差"])
        n = len(df)
        if n == 0:
            continue
        rows.append({
            "场景": scene,
            "样本数": n,
            "扣分完全准确数": int((df["扣分完全准确"] == 1).sum()),
            "扣分完全准确率": float((df["扣分完全准确"] == 1).mean()),
            "扣分1分内准确率": float((df["扣分绝对误差"] <= 1).mean()),
            "扣分2分内准确率": float((df["扣分绝对误差"] <= 2).mean()),
            "扣分MAE": float(df["扣分绝对误差"].mean()),
            "扣分RMSE": rmse(df["扣分误差"]),
            "平均扣分误差": float(df["扣分误差"].mean()),
            "AI扣分偏多数": int((df["扣分误差"] > 0).sum()),
            "AI扣分偏少数": int((df["扣分误差"] < 0).sum()),
            "平均实际扣分": float(df["实际扣分"].mean()),
            "平均预测扣分": float(df["预测扣分"].mean()),
        })
    return pd.DataFrame(rows)


def make_label_summary(label_detail_df: pd.DataFrame):
    if label_detail_df.empty:
        return pd.DataFrame()

    rows = []
    grouped = label_detail_df.dropna(subset=["实际标签", "预测标签"]).groupby(["场景", "标签序号", "标签名称"], as_index=False)

    for _, df in grouped:
        scene = df["场景"].iloc[0]
        label_idx = int(df["标签序号"].iloc[0])
        label_name = df["标签名称"].iloc[0]
        true = df["实际标签"].astype(int)
        pred = df["预测标签"].astype(int)
        prob = pd.to_numeric(df["预测概率"], errors="coerce")

        tp = int(((true == 1) & (pred == 1)).sum())
        tn = int(((true == 0) & (pred == 0)).sum())
        fp = int(((true == 0) & (pred == 1)).sum())
        fn = int(((true == 1) & (pred == 0)).sum())
        n = len(df)
        precision = tp / (tp + fp) if (tp + fp) > 0 else np.nan
        recall = tp / (tp + fn) if (tp + fn) > 0 else np.nan
        f1 = 2 * precision * recall / (precision + recall) if not pd.isna(precision) and not pd.isna(recall) and (precision + recall) > 0 else np.nan
        acc = (tp + tn) / n if n > 0 else np.nan

        rows.append({
            "场景": scene,
            "标签序号": label_idx,
            "标签名称": label_name,
            "样本数": n,
            "正样本数": int((true == 1).sum()),
            "负样本数": int((true == 0).sum()),
            "TP": tp,
            "TN": tn,
            "FP误扣": fp,
            "FN漏扣": fn,
            "Accuracy": acc,
            "Precision": precision,
            "Recall": recall,
            "F1": f1,
            "平均预测概率": float(prob.mean()) if len(prob.dropna()) > 0 else np.nan,
        })

    return pd.DataFrame(rows)


# =========================
# 7. Excel 导出与格式美化
# =========================

def safe_sheet_name(name):
    name = re.sub(r"[\\/*?:\[\]]", "_", str(name))
    return name[:31]


def write_excel_report(output_path: Path, score_summary_df, deduct_summary_df, label_summary_df, score_detail_df, image_detail_df, label_detail_df, skipped_df):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        score_summary_df.to_excel(writer, sheet_name="场景得分准确率", index=False)
        deduct_summary_df.to_excel(writer, sheet_name="单图扣分准确率", index=False)
        label_summary_df.to_excel(writer, sheet_name="标签识别统计", index=False)
        score_detail_df.to_excel(writer, sheet_name="每户得分明细", index=False)
        image_detail_df.to_excel(writer, sheet_name="每张图片明细", index=False)
        label_detail_df.to_excel(writer, sheet_name="标签概率明细", index=False)
        skipped_df.to_excel(writer, sheet_name="跳过样本", index=False)

        wb = writer.book

        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.chart import BarChart, Reference
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")
                    if isinstance(cell.value, float):
                        if 0 <= cell.value <= 1 and ("率" in str(ws.cell(1, cell.column).value) or str(ws.cell(1, cell.column).value) in ["Accuracy", "Precision", "Recall", "F1"]):
                            cell.number_format = "0.00%"
                        else:
                            cell.number_format = "0.00"

            # 自动列宽，上限防止过宽
            for col_cells in ws.columns:
                col_letter = get_column_letter(col_cells[0].column)
                max_len = 0
                for cell in col_cells:
                    text = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(text))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)

            ws.auto_filter.ref = ws.dimensions

        # 在“场景得分准确率”中加两个图：准确率与 MAE
        if "场景得分准确率" in wb.sheetnames and len(score_summary_df) > 0:
            ws = wb["场景得分准确率"]
            max_row = ws.max_row

            # 得分准确率柱状图
            chart1 = BarChart()
            chart1.title = "各场景得分完全准确率"
            chart1.y_axis.title = "准确率"
            chart1.x_axis.title = "项目"
            data = Reference(ws, min_col=4, min_row=1, max_row=max_row)  # 得分准确率
            cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)
            chart1.height = 7
            chart1.width = 15
            ws.add_chart(chart1, "T2")

            # MAE柱状图
            chart2 = BarChart()
            chart2.title = "各场景 MAE 平均绝对误差"
            chart2.y_axis.title = "MAE"
            chart2.x_axis.title = "项目"
            data2 = Reference(ws, min_col=9, min_row=1, max_row=max_row)  # MAE
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.height = 7
            chart2.width = 15
            ws.add_chart(chart2, "T18")

        # 在“单图扣分准确率”中加扣分准确率图
        if "单图扣分准确率" in wb.sheetnames and len(deduct_summary_df) > 0:
            ws = wb["单图扣分准确率"]
            max_row = ws.max_row
            chart = BarChart()
            chart.title = "各图片场景扣分完全准确率"
            chart.y_axis.title = "准确率"
            chart.x_axis.title = "场景"
            data = Reference(ws, min_col=4, min_row=1, max_row=max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 7
            chart.width = 15
            ws.add_chart(chart, "N2")


def print_console_summary(score_summary_df):
    print("\n========== 场景得分准确率汇总 ==========")
    if score_summary_df.empty:
        print("没有可统计的数据")
        return
    display_cols = ["项目", "样本数", "得分准确率", "1分内准确率", "2分内准确率", "MAE平均绝对误差", "平均分差"]
    print(score_summary_df[display_cols].to_string(index=False))


# =========================
# 8. 主函数
# =========================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=str, default="data/raw", help="农户文件夹根目录，例如 data/raw")
    parser.add_argument("--output", type=str, default="outputs/scene_score_accuracy_report.xlsx", help="输出 Excel 路径")
    args = parser.parse_args()

    root_dir = Path(args.root)
    output_path = Path(args.output)

    if not root_dir.exists():
        raise FileNotFoundError(f"找不到根目录: {root_dir}")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print("当前设备:", device)

    score_detail_df, image_detail_df, label_detail_df, skipped_df = collect_all_results(root_dir, device)

    if score_detail_df.empty:
        print("没有得到任何有效评分结果，请检查 data/raw 结构、CSV 和模型文件。")
        if not skipped_df.empty:
            print("\n跳过样本：")
            print(skipped_df.to_string(index=False))
        return

    score_summary_df = make_score_summary(score_detail_df)
    deduct_summary_df = make_deduct_summary(image_detail_df)
    label_summary_df = make_label_summary(label_detail_df)

    write_excel_report(
        output_path=output_path,
        score_summary_df=score_summary_df,
        deduct_summary_df=deduct_summary_df,
        label_summary_df=label_summary_df,
        score_detail_df=score_detail_df,
        image_detail_df=image_detail_df,
        label_detail_df=label_detail_df,
        skipped_df=skipped_df,
    )

    print_console_summary(score_summary_df)
    print("\nExcel 报告已生成:", output_path)


if __name__ == "__main__":
    main()
