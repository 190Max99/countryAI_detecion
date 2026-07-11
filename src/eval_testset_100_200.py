"""
测试集批量评分脚本：统计编号 100~200 的农户文件夹评分结果，并生成 Excel。

默认依赖当前项目中的：
    src/ui_folder_score_annotated_gradcam_multi.py

该模块需要提供：
    SCENE_CONFIGS
    prepare_label_df
    find_label_csv
    find_true_row
    calc_deduct_from_row
    find_image_in_folder
    predict_scene

运行方式：
    python -m src.eval_testset_100_200 \
        --root data/raw \
        --start 100 \
        --end 200 \
        --output outputs/testset_100_200_score_report.xlsx

输出 Excel 包含：
    1. 每户评分明细
    2. 场景评分统计
    3. 扣分项统计
    4. 测试集总体概览
    5. 跳过记录
"""

import argparse
import math
from pathlib import Path

import numpy as np
import pandas as pd
import torch


# ----------------------------------------------------------------------
# 复用现有 UI/预测模块中的配置与模型加载逻辑。
# 若你的文件名不是 ui_folder_score_annotated_gradcam_multi.py，
# 只需要修改下面这一行 import。
# ----------------------------------------------------------------------
from src.ui_folder_score_annotated_gradcam_multi import (
    SCENE_CONFIGS,
    prepare_label_df,
    find_label_csv,
    find_true_row,
    calc_deduct_from_row,
    find_image_in_folder,
    predict_scene,
)


SINGLE_SCENES = ["室内", "庭院", "厕所", "化粪池", "房前屋后"]
SCORE_PROJECTS = ["室内", "庭院", "厕所及化粪池", "房前屋后", "总分"]


def safe_float(value):
    """尽量把值转成 float，失败时返回 NaN。"""
    try:
        if pd.isna(value):
            return np.nan
        return float(value)
    except Exception:
        return np.nan


def calc_score(base_score, deduct):
    """根据满分和扣分计算得分。"""
    if pd.isna(deduct):
        return np.nan
    return max(0.0, float(base_score) - float(deduct))


def calc_error(true_score, pred_score):
    """分差 = AI预测得分 - 人工实际得分。"""
    if pd.isna(true_score) or pd.isna(pred_score):
        return np.nan
    return float(pred_score) - float(true_score)


def calc_binary_metrics(tp, tn, fp, fn):
    """根据混淆矩阵统计二分类指标。"""
    total = tp + tn + fp + fn

    accuracy = (tp + tn) / total if total > 0 else np.nan
    precision = tp / (tp + fp) if (tp + fp) > 0 else np.nan
    recall = tp / (tp + fn) if (tp + fn) > 0 else np.nan

    if not pd.isna(precision) and not pd.isna(recall) and (precision + recall) > 0:
        f1 = 2 * precision * recall / (precision + recall)
    else:
        f1 = np.nan

    mismatch_rate = (fp + fn) / total if total > 0 else np.nan
    false_positive_rate = fp / total if total > 0 else np.nan
    false_negative_rate = fn / total if total > 0 else np.nan

    return {
        "Accuracy": accuracy,
        "Precision": precision,
        "Recall": recall,
        "F1": f1,
        "不一致率": mismatch_rate,
        "误扣率": false_positive_rate,
        "漏扣率": false_negative_rate,
    }


def evaluate_one_folder(folder: Path, device):
    """
    评价一个农户文件夹。

    返回：
        house_score_rows：该户各评分项目明细
        label_rows：该户每个扣分项预测明细
        errors：跳过或错误信息
    """
    house_id = folder.name
    errors = []
    house_score_rows = []
    label_rows = []

    label_csv = find_label_csv(folder)

    if label_csv is None:
        errors.append({
            "house_id": house_id,
            "场景": "整户",
            "原因": "文件夹中未找到人工标签 CSV",
            "路径": str(folder),
        })
        return house_score_rows, label_rows, errors

    try:
        label_df = prepare_label_df(label_csv)
    except Exception as exc:
        errors.append({
            "house_id": house_id,
            "场景": "整户",
            "原因": f"人工标签 CSV 读取失败：{exc}",
            "路径": str(label_csv),
        })
        return house_score_rows, label_rows, errors

    scene_results = {}

    for scene_key in SINGLE_SCENES:
        true_row = find_true_row(label_df, scene_key)
        image_path = find_image_in_folder(folder, scene_key)

        if true_row is None:
            errors.append({
                "house_id": house_id,
                "场景": scene_key,
                "原因": "人工标签 CSV 中未找到该场景",
                "路径": str(label_csv),
            })
            continue

        if image_path is None:
            errors.append({
                "house_id": house_id,
                "场景": scene_key,
                "原因": "未找到对应场景图片",
                "路径": str(folder),
            })
            continue

        try:
            pred_info = predict_scene(scene_key, image_path, device)
        except Exception as exc:
            errors.append({
                "house_id": house_id,
                "场景": scene_key,
                "原因": f"模型预测失败：{exc}",
                "路径": str(image_path),
            })
            continue

        true_deduct = float(calc_deduct_from_row(true_row, scene_key))
        pred_deduct = float(pred_info["pred_deduct"])

        scene_results[scene_key] = {
            "true_deduct": true_deduct,
            "pred_deduct": pred_deduct,
            "image_path": str(image_path),
        }

        # 每个扣分项明细
        cfg = SCENE_CONFIGS[scene_key]
        true_labels = []

        for col in cfg["label_cols"]:
            value = 0
            if col in true_row.index and not pd.isna(true_row[col]):
                try:
                    value = int(float(true_row[col]))
                except Exception:
                    value = 0
            true_labels.append(value)

        pred_labels = np.asarray(pred_info["pred_labels"]).astype(int)
        probs = np.asarray(pred_info["probs"], dtype=float)
        thresholds = np.asarray(pred_info["thresholds"], dtype=float)
        deducts = list(pred_info["deducts"])
        label_names = list(pred_info["label_names"])

        count = min(
            len(true_labels),
            len(pred_labels),
            len(probs),
            len(thresholds),
            len(deducts),
            len(label_names),
        )

        for idx in range(count):
            true_label = int(true_labels[idx])
            pred_label = int(pred_labels[idx])

            if true_label == 1 and pred_label == 1:
                result_type = "正确扣分"
            elif true_label == 0 and pred_label == 0:
                result_type = "正确不扣"
            elif true_label == 0 and pred_label == 1:
                result_type = "误扣"
            else:
                result_type = "漏扣"

            label_rows.append({
                "house_id": house_id,
                "场景": scene_key,
                "标签序号": idx,
                "标签列": cfg["label_cols"][idx],
                "扣分项": label_names[idx],
                "扣分值": deducts[idx],
                "人工标签": true_label,
                "AI预测标签": pred_label,
                "是否一致": int(true_label == pred_label),
                "错误类型": result_type,
                "预测概率": float(probs[idx]),
                "判断阈值": float(thresholds[idx]),
                "图片路径": str(image_path),
            })

    # -------------------------------
    # 生成四个场景与总分明细
    # -------------------------------
    def add_score_row(project, full_score, true_deduct, pred_deduct):
        true_score = calc_score(full_score, true_deduct)
        pred_score = calc_score(full_score, pred_deduct)
        error = calc_error(true_score, pred_score)

        house_score_rows.append({
            "house_id": house_id,
            "项目": project,
            "满分": full_score,
            "人工扣分": true_deduct,
            "AI预测扣分": pred_deduct,
            "人工实际得分": true_score,
            "AI预测得分": pred_score,
            "分差": error,
            "绝对误差": abs(error) if not pd.isna(error) else np.nan,
            "是否完全准确": int(error == 0) if not pd.isna(error) else np.nan,
            "是否1分内": int(abs(error) <= 1) if not pd.isna(error) else np.nan,
            "是否2分内": int(abs(error) <= 2) if not pd.isna(error) else np.nan,
            "是否3分内": int(abs(error) <= 3) if not pd.isna(error) else np.nan,
            "是否5分内": int(abs(error) <= 5) if not pd.isna(error) else np.nan,
        })

    if "室内" in scene_results:
        add_score_row(
            "室内",
            10,
            scene_results["室内"]["true_deduct"],
            scene_results["室内"]["pred_deduct"],
        )

    if "庭院" in scene_results:
        add_score_row(
            "庭院",
            30,
            scene_results["庭院"]["true_deduct"],
            scene_results["庭院"]["pred_deduct"],
        )

    if "厕所" in scene_results and "化粪池" in scene_results:
        true_deduct = (
            scene_results["厕所"]["true_deduct"]
            + scene_results["化粪池"]["true_deduct"]
        )
        pred_deduct = (
            scene_results["厕所"]["pred_deduct"]
            + scene_results["化粪池"]["pred_deduct"]
        )

        add_score_row(
            "厕所及化粪池",
            10,
            true_deduct,
            pred_deduct,
        )

    if "房前屋后" in scene_results:
        add_score_row(
            "房前屋后",
            10,
            scene_results["房前屋后"]["true_deduct"],
            scene_results["房前屋后"]["pred_deduct"],
        )

    # 只有四个评分项目全部存在时才统计总分
    current_projects = {
        row["项目"]: row for row in house_score_rows
        if row["项目"] != "总分"
    }

    required_projects = {"室内", "庭院", "厕所及化粪池", "房前屋后"}

    if required_projects.issubset(current_projects.keys()):
        true_total = sum(
            current_projects[name]["人工实际得分"]
            for name in required_projects
        )
        pred_total = sum(
            current_projects[name]["AI预测得分"]
            for name in required_projects
        )

        true_total_deduct = 60 - true_total
        pred_total_deduct = 60 - pred_total

        add_score_row(
            "总分",
            60,
            true_total_deduct,
            pred_total_deduct,
        )
    else:
        missing = sorted(required_projects - set(current_projects.keys()))
        errors.append({
            "house_id": house_id,
            "场景": "总分",
            "原因": "缺少评分项目，无法计算整户总分：" + "、".join(missing),
            "路径": str(folder),
        })

    return house_score_rows, label_rows, errors


def build_scene_summary(score_df: pd.DataFrame):
    """按项目统计得分准确率和误差。"""
    rows = []

    if score_df.empty:
        return pd.DataFrame()

    for project in SCORE_PROJECTS:
        part = score_df[score_df["项目"] == project].copy()
        part = part.dropna(
            subset=["人工实际得分", "AI预测得分", "分差", "绝对误差"]
        )

        if part.empty:
            continue

        sample_count = len(part)
        exact_count = int((part["绝对误差"] == 0).sum())

        rows.append({
            "项目": project,
            "样本数": sample_count,
            "得分完全准确数": exact_count,
            "得分准确率": exact_count / sample_count,
            "1分内准确率": float((part["绝对误差"] <= 1).mean()),
            "2分内准确率": float((part["绝对误差"] <= 2).mean()),
            "3分内准确率": float((part["绝对误差"] <= 3).mean()),
            "5分内准确率": float((part["绝对误差"] <= 5).mean()),
            "MAE平均绝对误差": float(part["绝对误差"].mean()),
            "RMSE均方根误差": float(
                math.sqrt(np.mean(np.square(part["分差"])))
            ),
            "平均分差": float(part["分差"].mean()),
            "AI给分偏高数": int((part["分差"] > 0).sum()),
            "AI给分偏低数": int((part["分差"] < 0).sum()),
            "最高实际得分": float(part["人工实际得分"].max()),
            "最低实际得分": float(part["人工实际得分"].min()),
            "平均实际得分": float(part["人工实际得分"].mean()),
            "平均预测得分": float(part["AI预测得分"].mean()),
        })

    return pd.DataFrame(rows)


def build_label_summary(label_df: pd.DataFrame):
    """按场景和扣分项统计一致率、误扣率、漏扣率等指标。"""
    rows = []

    if label_df.empty:
        return pd.DataFrame()

    group_cols = ["场景", "标签序号", "标签列", "扣分项", "扣分值"]

    for keys, part in label_df.groupby(group_cols, dropna=False):
        scene, label_idx, label_col, label_name, deduct = keys

        y_true = part["人工标签"].astype(int).to_numpy()
        y_pred = part["AI预测标签"].astype(int).to_numpy()

        tp = int(((y_true == 1) & (y_pred == 1)).sum())
        tn = int(((y_true == 0) & (y_pred == 0)).sum())
        fp = int(((y_true == 0) & (y_pred == 1)).sum())
        fn = int(((y_true == 1) & (y_pred == 0)).sum())

        metrics = calc_binary_metrics(tp, tn, fp, fn)

        rows.append({
            "场景": scene,
            "标签序号": label_idx,
            "标签列": label_col,
            "扣分项": label_name,
            "扣分值": deduct,
            "有效样本数": len(part),
            "人工扣分次数": int((y_true == 1).sum()),
            "AI扣分次数": int((y_pred == 1).sum()),
            "一致次数": tp + tn,
            "不一致次数": fp + fn,
            "误扣次数": fp,
            "漏扣次数": fn,
            "正确扣分TP": tp,
            "正确不扣TN": tn,
            "平均预测概率": float(part["预测概率"].mean()),
            "阈值": float(part["判断阈值"].iloc[0]),
            **metrics,
        })

    result = pd.DataFrame(rows)

    if not result.empty:
        scene_order = {
            "室内": 0,
            "庭院": 1,
            "厕所": 2,
            "化粪池": 3,
            "房前屋后": 4,
        }
        result["_scene_order"] = result["场景"].map(scene_order).fillna(99)
        result = result.sort_values(
            ["_scene_order", "标签序号"]
        ).drop(columns=["_scene_order"])

    return result


def build_overview(
    start,
    end,
    existing_folders,
    score_df,
    label_df,
    skipped_df,
):
    """生成测试集总体概览。"""
    expected_count = end - start + 1

    total_rows = score_df[score_df["项目"] == "总分"] if not score_df.empty else pd.DataFrame()
    exact_total = (
        int((total_rows["绝对误差"] == 0).sum())
        if not total_rows.empty else 0
    )

    overview = [
        ["测试集编号范围", f"{start}~{end}"],
        ["理论文件夹数量", expected_count],
        ["实际存在文件夹数量", len(existing_folders)],
        ["成功得到总分的农户数", len(total_rows)],
        ["总分完全准确农户数", exact_total],
        [
            "总分完全准确率",
            exact_total / len(total_rows) if len(total_rows) > 0 else np.nan,
        ],
        [
            "总分2分内准确率",
            float((total_rows["绝对误差"] <= 2).mean())
            if len(total_rows) > 0 else np.nan,
        ],
        [
            "总分MAE",
            float(total_rows["绝对误差"].mean())
            if len(total_rows) > 0 else np.nan,
        ],
        ["扣分项有效预测记录数", len(label_df)],
        ["跳过或异常记录数", len(skipped_df)],
        ["运行设备", "CUDA" if torch.cuda.is_available() else "CPU"],
    ]

    return pd.DataFrame(overview, columns=["指标", "结果"])


def style_excel(output_path: Path):
    """使用 openpyxl 对生成的 Excel 做基础格式美化。"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
    except ImportError:
        print("未安装 openpyxl，Excel 已生成，但未进行格式美化。")
        return

    wb = load_workbook(output_path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    percent_keywords = [
        "准确率", "不一致率", "误扣率", "漏扣率",
        "Accuracy", "Precision", "Recall", "F1",
    ]

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        # 自动设置列宽，并限制最大宽度
        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter
            max_len = 0

            for cell in column_cells:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))

            ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 42)

        # 百分比格式
        headers = {
            cell.column: str(cell.value)
            for cell in ws[1]
            if cell.value is not None
        }

        for col_idx, header in headers.items():
            if any(key in header for key in percent_keywords):
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"

        # 对误差列做色阶
        for col_idx, header in headers.items():
            if header in {"分差", "绝对误差", "MAE平均绝对误差", "RMSE均方根误差"}:
                if ws.max_row >= 2:
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    ws.conditional_formatting.add(
                        f"{col_letter}2:{col_letter}{ws.max_row}",
                        ColorScaleRule(
                            start_type="min",
                            start_color="63BE7B",
                            mid_type="percentile",
                            mid_value=50,
                            mid_color="FFEB84",
                            end_type="max",
                            end_color="F8696B",
                        ),
                    )

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="统计编号 100~200 测试集的场景评分与扣分项准确率"
    )

    parser.add_argument(
        "--root",
        type=str,
        default="data/raw",
        help="农户文件夹根目录，默认 data/raw",
    )

    parser.add_argument(
        "--start",
        type=int,
        default=100,
        help="测试集起始编号，默认 100",
    )

    parser.add_argument(
        "--end",
        type=int,
        default=200,
        help="测试集结束编号，默认 200",
    )

    parser.add_argument(
        "--output",
        type=str,
        default="outputs/testset_100_200_score_report.xlsx",
        help="Excel 输出路径",
    )

    args = parser.parse_args()

    if args.start > args.end:
        raise ValueError("--start 不能大于 --end")

    root_dir = Path(args.root)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not root_dir.exists():
        raise FileNotFoundError(f"测试集根目录不存在：{root_dir}")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    print("=" * 70)
    print("测试集批量评分")
    print(f"文件夹范围：{args.start} ~ {args.end}")
    print(f"测试集根目录：{root_dir}")
    print(f"运行设备：{device}")
    print("=" * 70)

    all_score_rows = []
    all_label_rows = []
    all_errors = []
    existing_folders = []

    for house_number in range(args.start, args.end + 1):
        folder = root_dir / str(house_number)

        if not folder.exists() or not folder.is_dir():
            all_errors.append({
                "house_id": str(house_number),
                "场景": "整户",
                "原因": "测试集文件夹不存在",
                "路径": str(folder),
            })
            print(f"[跳过] {house_number}：文件夹不存在")
            continue

        existing_folders.append(folder)
        print(f"[测试] {house_number}")

        score_rows, label_rows, errors = evaluate_one_folder(
            folder=folder,
            device=device,
        )

        all_score_rows.extend(score_rows)
        all_label_rows.extend(label_rows)
        all_errors.extend(errors)

    score_df = pd.DataFrame(all_score_rows)
    label_df = pd.DataFrame(all_label_rows)
    skipped_df = pd.DataFrame(all_errors)

    scene_summary_df = build_scene_summary(score_df)
    label_summary_df = build_label_summary(label_df)

    overview_df = build_overview(
        start=args.start,
        end=args.end,
        existing_folders=existing_folders,
        score_df=score_df,
        label_df=label_df,
        skipped_df=skipped_df,
    )

    # 排序，让 Excel 更容易查看
    if not score_df.empty:
        score_df["_house_num"] = pd.to_numeric(
            score_df["house_id"], errors="coerce"
        )
        score_order = {name: idx for idx, name in enumerate(SCORE_PROJECTS)}
        score_df["_project_order"] = score_df["项目"].map(score_order)
        score_df = score_df.sort_values(
            ["_house_num", "_project_order"]
        ).drop(columns=["_house_num", "_project_order"])

    if not label_df.empty:
        label_df["_house_num"] = pd.to_numeric(
            label_df["house_id"], errors="coerce"
        )
        label_df = label_df.sort_values(
            ["_house_num", "场景", "标签序号"]
        ).drop(columns=["_house_num"])

    if skipped_df.empty:
        skipped_df = pd.DataFrame(
            columns=["house_id", "场景", "原因", "路径"]
        )

    print("\n正在生成 Excel 报告……")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overview_df.to_excel(
            writer,
            sheet_name="测试集总体概览",
            index=False,
        )

        scene_summary_df.to_excel(
            writer,
            sheet_name="场景评分统计",
            index=False,
        )

        score_df.to_excel(
            writer,
            sheet_name="每户评分明细",
            index=False,
        )

        label_summary_df.to_excel(
            writer,
            sheet_name="扣分项统计",
            index=False,
        )

        label_df.to_excel(
            writer,
            sheet_name="每户扣分项明细",
            index=False,
        )

        skipped_df.to_excel(
            writer,
            sheet_name="跳过记录",
            index=False,
        )

    style_excel(output_path)

    print("\n测试完成")
    print(f"实际存在文件夹数：{len(existing_folders)}")
    print(
        "成功统计总分农户数：",
        len(score_df[score_df["项目"] == "总分"])
        if not score_df.empty else 0,
    )
    print(f"扣分项预测记录数：{len(label_df)}")
    print(f"跳过或异常记录数：{len(skipped_df)}")
    print(f"Excel 报告：{output_path}")


if __name__ == "__main__":
    main()
