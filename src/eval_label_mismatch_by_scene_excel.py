"""
按“扣分项/问题”评价 AI 与人工是否一致，并生成 Excel 报告。

作用：
1. 批量扫描 data/raw 下每个农户文件夹；
2. 读取每户 CSV 中的人工标签；
3. 调用每个场景的 ResNet18 / ResNet18+CBAM 模型预测标签；
4. 对每个扣分项统计：不一致率、误扣率、漏扣率、Precision、Recall、F1；
5. 输出 Excel，便于分析“庭院 B5 地面垃圾”等具体问题的模型表现。

用法：
python -m src.eval_label_mismatch_by_scene_excel --root data/raw --output outputs/label_mismatch_report.xlsx

依赖：
pip install pandas numpy pillow torch torchvision openpyxl
"""

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from PIL import Image

import torch
import torch.nn as nn
from torchvision import models, transforms


# =========================
# 1. 场景与标签配置
# =========================

SCENE_CONFIGS = {
    "室内": {
        "aliases": ["室内"],
        "model_path": "models/indoor_resnet18.pth",
        "fallback_model_paths": [],
        "label_cols": [f"label_{i}" for i in range(10)],
        "label_names": [
            "A0_室内格局杂乱无章",
            "A1_室内家具摆放杂乱无章",
            "A2_室内生活用品杂乱无章",
            "A3_鸡鸭进入屋内共居",
            "A4_室内地面存在鸡鸭粪污",
            "A5_鸡跳在室内桌子上",
            "A6_室内地面垃圾乱丢现象严重",
            "A7_室内桌面沙发表面乱堆乱摆",
            "A8_室内墙面污迹不堪",
            "A9_室内其他脏乱情况",
        ],
        "deducts": [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
        "thresholds": [0.55, 0.55, 0.55, 0.45, 0.45, 0.45, 0.50, 0.50, 0.50, 0.65],
    },
    "庭院": {
        "aliases": ["庭院"],
        "model_path": "models/courtyard_resnet18_cbam.pth",
        "fallback_model_paths": ["models/courtyard_resnet18.pth"],
        "label_cols": [f"label_{i}" for i in range(12)],
        "label_names": [
            "B0_庭院内生产工具杂乱无章",
            "B1_庭院内交通用具杂乱无章",
            "B2_庭院内其他情况杂乱无章",
            "B3_鸡鸭进入庭院乱跑",
            "B4_庭院内鸡粪鸭粪满地",
            "B5_庭院内地面垃圾乱丢现象严重",
            "B6_庭院内柴草堆码无序",
            "B7_庭院内柴草堆码不整齐",
            "B8_庭院内房屋立面乱挂乱画",
            "B9_庭院内搭建棚库破败物品堆码杂乱不堪",
            "B10_庭院污水横流",
            "B11_庭院内其他情况",
        ],
        "deducts": [3, 3, 2, 2, 2, 5, 1, 2, 2, 5, 2, 1],
        "thresholds": [0.55, 0.55, 0.60, 0.45, 0.45, 0.50, 0.55, 0.55, 0.55, 0.55, 0.50, 0.65],
    },
    "厕所": {
        "aliases": ["厕所", "厕屋"],
        "model_path": "models/toilet_resnet18.pth",
        "fallback_model_paths": [],
        "label_cols": [f"label_{i}" for i in range(2)],
        "label_names": [
            "C0_厕屋脏乱",
            "C1_厕屋功能配备不齐全",
        ],
        "deducts": [2, 3],
        "thresholds": [0.50, 0.50],
    },
    "化粪池": {
        "aliases": ["化粪池"],
        "model_path": "models/septic_resnet18.pth",
        "fallback_model_paths": [],
        "label_cols": [f"label_{i}" for i in range(3)],
        "label_names": [
            "C2_化粪池盖板挪开未关闭取粪口未关闭",
            "C3_化粪池粪污溢流",
            "C4_厕所周围其他情况",
        ],
        "deducts": [2, 2, 1],
        "thresholds": [0.50, 0.50, 0.65],
    },
    "房前屋后": {
        "aliases": ["房前屋后", "房屋前后", "房前屋后及2侧", "房前屋后及两侧", "屋后", "两侧", "2侧"],
        "model_path": "models/outside_resnet18.pth",
        "fallback_model_paths": [],
        "label_cols": [f"label_{i}" for i in range(5)],
        "label_names": [
            "D0_房屋旁柴草堆码乱堆不整齐",
            "D1_房屋周身存在污水横流现象",
            "D2_房屋周身瓜果棚架破败不堪",
            "D3_房屋周身鸡鸭棚圈破败不堪脏臭",
            "D4_房屋周身其他情况",
        ],
        "deducts": [3, 2, 2, 2, 1],
        "thresholds": [0.55, 0.50, 0.50, 0.50, 0.65],
    },
}

IMAGE_EXTS = [".jpg", ".jpeg", ".png", ".bmp", ".webp"]
MODEL_CACHE = {}


# =========================
# 2. CBAM 结构，用于加载庭院 CBAM 模型
# =========================

class ChannelAttention(nn.Module):
    def __init__(self, in_channels, reduction=16):
        super().__init__()
        hidden_channels = max(in_channels // reduction, 1)
        self.avg_pool = nn.AdaptiveAvgPool2d(1)
        self.max_pool = nn.AdaptiveMaxPool2d(1)
        self.mlp = nn.Sequential(
            nn.Conv2d(in_channels, hidden_channels, kernel_size=1, bias=False),
            nn.ReLU(inplace=True),
            nn.Conv2d(hidden_channels, in_channels, kernel_size=1, bias=False),
        )
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg_out = self.mlp(self.avg_pool(x))
        max_out = self.mlp(self.max_pool(x))
        attention = self.sigmoid(avg_out + max_out)
        return x * attention


class SpatialAttention(nn.Module):
    def __init__(self, kernel_size=7):
        super().__init__()
        padding = kernel_size // 2
        self.conv = nn.Conv2d(2, 1, kernel_size=kernel_size, padding=padding, bias=False)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        avg_out = torch.mean(x, dim=1, keepdim=True)
        max_out, _ = torch.max(x, dim=1, keepdim=True)
        attention_input = torch.cat([avg_out, max_out], dim=1)
        attention = self.sigmoid(self.conv(attention_input))
        return x * attention


class CBAM(nn.Module):
    def __init__(self, in_channels, reduction=16, kernel_size=7):
        super().__init__()
        self.channel_attention = ChannelAttention(in_channels, reduction)
        self.spatial_attention = SpatialAttention(kernel_size)

    def forward(self, x):
        x = self.channel_attention(x)
        x = self.spatial_attention(x)
        return x


class ResNet18CBAM(nn.Module):
    def __init__(self, num_labels):
        super().__init__()
        base_model = models.resnet18(weights=None)
        self.conv1 = base_model.conv1
        self.bn1 = base_model.bn1
        self.relu = base_model.relu
        self.maxpool = base_model.maxpool
        self.layer1 = base_model.layer1
        self.layer2 = base_model.layer2
        self.layer3 = base_model.layer3
        self.layer4 = base_model.layer4
        self.cbam = CBAM(in_channels=512)
        self.avgpool = base_model.avgpool
        self.fc = nn.Linear(512, num_labels)

    def forward(self, x):
        x = self.conv1(x)
        x = self.bn1(x)
        x = self.relu(x)
        x = self.maxpool(x)
        x = self.layer1(x)
        x = self.layer2(x)
        x = self.layer3(x)
        x = self.layer4(x)
        x = self.cbam(x)
        x = self.avgpool(x)
        x = torch.flatten(x, 1)
        x = self.fc(x)
        return x


# =========================
# 3. 工具函数
# =========================

def normalize_text(value):
    if pd.isna(value):
        return ""
    text = str(value)
    text = text.replace("\ufeff", "")
    text = text.replace("\t", "")
    text = text.replace(" ", "")
    text = text.replace("\u3000", "")
    return text.strip()


def read_csv_safely(csv_path: Path):
    encodings = ["utf-8-sig", "gbk", "gb18030", "utf-8"]
    seps = [None, "\t", ","]
    best_df = None
    best_col_count = 0

    for enc in encodings:
        for sep in seps:
            try:
                if sep is None:
                    df = pd.read_csv(csv_path, encoding=enc, sep=None, engine="python")
                else:
                    df = pd.read_csv(csv_path, encoding=enc, sep=sep)
                df.columns = [normalize_text(c) for c in df.columns]
                if len(df.columns) > best_col_count:
                    best_df = df
                    best_col_count = len(df.columns)
                if len(df.columns) >= 3:
                    return df
            except Exception:
                continue

    if best_df is not None:
        return best_df
    raise RuntimeError(f"无法读取 CSV 文件: {csv_path}")


def find_label_csv(folder: Path):
    preferred = folder / f"{folder.name}.csv"
    if preferred.exists():
        return preferred

    csv_files = sorted(folder.glob("*.csv"))
    bad_keywords = ["result", "detail", "score", "预测", "结果", "simple", "ui", "report"]
    valid_files = []

    for file in csv_files:
        lower_name = file.name.lower()
        if any(k.lower() in lower_name for k in bad_keywords):
            continue
        valid_files.append(file)

    if valid_files:
        return valid_files[0]
    if csv_files:
        return csv_files[0]
    return None


def prepare_label_df(label_csv: Path):
    df = read_csv_safely(label_csv)

    if "scene" in df.columns:
        scene_col = "scene"
    elif "序号" in df.columns:
        scene_col = "序号"
    elif "类别" in df.columns:
        scene_col = "类别"
    else:
        scene_col = df.columns[0]

    df["scene_norm"] = df[scene_col].apply(normalize_text)
    return df


def find_true_row(label_df: pd.DataFrame, scene_key: str):
    aliases = SCENE_CONFIGS[scene_key]["aliases"]
    rows = label_df[label_df["scene_norm"].isin(aliases)]
    if len(rows) > 0:
        return rows.iloc[0]

    for _, row in label_df.iterrows():
        scene_text = str(row["scene_norm"])
        for alias in aliases:
            if alias in scene_text:
                return row
    return None


def get_label_value(row, col):
    if row is None:
        return 0
    if col not in row.index:
        return 0
    value = row[col]
    if pd.isna(value):
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def find_image_in_folder(folder: Path, scene_key: str):
    aliases = SCENE_CONFIGS[scene_key]["aliases"]
    candidates = []

    for file in folder.iterdir():
        if not file.is_file():
            continue
        if file.suffix.lower() not in IMAGE_EXTS:
            continue
        filename = normalize_text(file.name)
        for alias in aliases:
            if alias in filename:
                candidates.append(file)
                break

    if not candidates:
        return None
    return sorted(candidates, key=lambda p: p.name)[0]


def get_transform():
    return transforms.Compose([
        transforms.Resize((224, 224)),
        transforms.ToTensor(),
        transforms.Normalize(
            mean=[0.485, 0.456, 0.406],
            std=[0.229, 0.224, 0.225]
        )
    ])


def resolve_model_path(scene_key):
    cfg = SCENE_CONFIGS[scene_key]
    candidates = [cfg["model_path"]] + cfg.get("fallback_model_paths", [])
    for text in candidates:
        path = Path(text)
        if path.exists():
            return path
    raise FileNotFoundError(f"{scene_key} 模型不存在，已尝试: {candidates}")


def load_checkpoint(model_path, device):
    try:
        return torch.load(model_path, map_location=device, weights_only=False)
    except TypeError:
        return torch.load(model_path, map_location=device)


def detect_cbam_model(checkpoint):
    if checkpoint.get("model_type", "") == "resnet18_cbam":
        return True
    state_dict = checkpoint.get("model_state_dict", {})
    return any(key.startswith("cbam.") for key in state_dict.keys())


def build_plain_resnet18(num_labels):
    model = models.resnet18(weights=None)
    model.fc = nn.Linear(model.fc.in_features, num_labels)
    return model


def load_model_for_scene(scene_key, device):
    cache_key = (scene_key, str(device))
    if cache_key in MODEL_CACHE:
        return MODEL_CACHE[cache_key]

    cfg = SCENE_CONFIGS[scene_key]
    model_path = resolve_model_path(scene_key)
    checkpoint = load_checkpoint(model_path, device)

    label_names = checkpoint.get("label_names", cfg["label_names"])
    thresholds = np.array(checkpoint.get("thresholds", cfg["thresholds"]), dtype=float)
    deducts = checkpoint.get("deducts", cfg["deducts"])
    use_cbam = detect_cbam_model(checkpoint)

    if use_cbam:
        model = ResNet18CBAM(num_labels=len(label_names))
    else:
        model = build_plain_resnet18(num_labels=len(label_names))

    model.load_state_dict(checkpoint["model_state_dict"])
    model.to(device)
    model.eval()

    pack = {
        "model": model,
        "label_names": label_names,
        "thresholds": thresholds,
        "deducts": deducts,
        "model_path": str(model_path),
        "use_cbam": use_cbam,
    }
    MODEL_CACHE[cache_key] = pack
    return pack


def predict_scene(scene_key, image_path, device):
    pack = load_model_for_scene(scene_key, device)
    model = pack["model"]
    image = Image.open(image_path).convert("RGB")
    image_tensor = get_transform()(image).unsqueeze(0).to(device)

    with torch.no_grad():
        logits = model(image_tensor)
        probs = torch.sigmoid(logits)[0].cpu().numpy()

    pred_labels = (probs >= pack["thresholds"]).astype(int)
    return probs, pred_labels, pack


def safe_div(n, d):
    if d == 0:
        return np.nan
    return n / d


# =========================
# 4. 主评价逻辑
# =========================

def evaluate(root_dir: Path):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print("当前设备:", device)

    label_detail_rows = []
    skip_rows = []

    folders = [p for p in sorted(root_dir.iterdir(), key=lambda x: x.name) if p.is_dir()]

    for folder in folders:
        house_id = folder.name
        label_csv = find_label_csv(folder)

        if label_csv is None:
            skip_rows.append({"house_id": house_id, "场景": "全部", "原因": "缺少人工标签 CSV"})
            continue

        try:
            label_df = prepare_label_df(label_csv)
        except Exception as e:
            skip_rows.append({"house_id": house_id, "场景": "全部", "原因": f"CSV 读取失败: {e}"})
            continue

        for scene_key, cfg in SCENE_CONFIGS.items():
            true_row = find_true_row(label_df, scene_key)
            if true_row is None:
                skip_rows.append({"house_id": house_id, "场景": scene_key, "原因": "CSV 中缺少该场景人工标签"})
                continue

            image_path = find_image_in_folder(folder, scene_key)
            if image_path is None:
                skip_rows.append({"house_id": house_id, "场景": scene_key, "原因": "缺少该场景图片"})
                continue

            try:
                probs, pred_labels, pack = predict_scene(scene_key, image_path, device)
            except Exception as e:
                skip_rows.append({"house_id": house_id, "场景": scene_key, "原因": f"模型预测失败: {e}"})
                continue

            label_names = pack["label_names"]
            thresholds = pack["thresholds"]
            deducts = pack["deducts"]
            label_cols = cfg["label_cols"]

            for i, col in enumerate(label_cols):
                true_label = get_label_value(true_row, col)
                pred_label = int(pred_labels[i])
                prob = float(probs[i])
                threshold = float(thresholds[i])
                consistent = int(true_label == pred_label)

                if true_label == 1 and pred_label == 1:
                    error_type = "正确扣分"
                elif true_label == 0 and pred_label == 0:
                    error_type = "正确不扣"
                elif true_label == 0 and pred_label == 1:
                    error_type = "误扣_AI扣人工未扣"
                else:
                    error_type = "漏扣_人工扣AI未扣"

                label_detail_rows.append({
                    "house_id": house_id,
                    "场景": scene_key,
                    "标签列": col,
                    "扣分项": label_names[i],
                    "扣分值": deducts[i],
                    "人工是否扣分": true_label,
                    "AI是否扣分": pred_label,
                    "是否一致": consistent,
                    "不一致": 1 - consistent,
                    "错误类型": error_type,
                    "预测概率": prob,
                    "阈值": threshold,
                    "图片路径": str(image_path),
                    "人工CSV": str(label_csv),
                })

    detail_df = pd.DataFrame(label_detail_rows)
    skip_df = pd.DataFrame(skip_rows)

    if len(detail_df) == 0:
        return detail_df, pd.DataFrame(), pd.DataFrame(), skip_df

    summary_rows = []

    group_cols = ["场景", "标签列", "扣分项", "扣分值"]
    for group_key, group in detail_df.groupby(group_cols, dropna=False):
        scene_key, label_col, label_name, deduct = group_key

        n = len(group)
        true_1 = int((group["人工是否扣分"] == 1).sum())
        true_0 = int((group["人工是否扣分"] == 0).sum())
        pred_1 = int((group["AI是否扣分"] == 1).sum())
        pred_0 = int((group["AI是否扣分"] == 0).sum())

        tp = int(((group["人工是否扣分"] == 1) & (group["AI是否扣分"] == 1)).sum())
        tn = int(((group["人工是否扣分"] == 0) & (group["AI是否扣分"] == 0)).sum())
        fp = int(((group["人工是否扣分"] == 0) & (group["AI是否扣分"] == 1)).sum())
        fn = int(((group["人工是否扣分"] == 1) & (group["AI是否扣分"] == 0)).sum())

        mismatch = fp + fn

        accuracy = safe_div(tp + tn, n)
        mismatch_rate = safe_div(mismatch, n)
        false_positive_rate_total = safe_div(fp, n)
        false_negative_rate_total = safe_div(fn, n)
        false_positive_rate_condition = safe_div(fp, true_0)
        false_negative_rate_condition = safe_div(fn, true_1)

        precision = safe_div(tp, tp + fp)
        recall = safe_div(tp, tp + fn)
        f1 = safe_div(2 * precision * recall, precision + recall) if not (pd.isna(precision) or pd.isna(recall)) else np.nan

        mean_prob = group["预测概率"].mean()
        mean_prob_true1 = group.loc[group["人工是否扣分"] == 1, "预测概率"].mean()
        mean_prob_true0 = group.loc[group["人工是否扣分"] == 0, "预测概率"].mean()
        threshold = group["阈值"].iloc[0]

        summary_rows.append({
            "场景": scene_key,
            "标签列": label_col,
            "扣分项": label_name,
            "扣分值": deduct,
            "有效样本数": n,
            "人工扣分次数": true_1,
            "人工不扣次数": true_0,
            "AI扣分次数": pred_1,
            "AI不扣次数": pred_0,
            "一致次数": tp + tn,
            "不一致次数": mismatch,
            "不一致率": mismatch_rate,
            "标签准确率": accuracy,
            "误扣次数_AI扣人工未扣": fp,
            "误扣率_占全部样本": false_positive_rate_total,
            "误扣率_占人工不扣样本": false_positive_rate_condition,
            "漏扣次数_人工扣AI未扣": fn,
            "漏扣率_占全部样本": false_negative_rate_total,
            "漏扣率_占人工扣分样本": false_negative_rate_condition,
            "Precision_扣分精确率": precision,
            "Recall_扣分召回率": recall,
            "F1": f1,
            "平均预测概率": mean_prob,
            "人工扣分样本平均概率": mean_prob_true1,
            "人工不扣样本平均概率": mean_prob_true0,
            "阈值": threshold,
        })

    label_summary_df = pd.DataFrame(summary_rows)

    # 场景级汇总：看每个场景所有扣分项整体不一致情况
    scene_rows = []
    for scene_key, group in detail_df.groupby("场景"):
        n = len(group)
        mismatch = int(group["不一致"].sum())
        fp = int((group["错误类型"] == "误扣_AI扣人工未扣").sum())
        fn = int((group["错误类型"] == "漏扣_人工扣AI未扣").sum())
        scene_rows.append({
            "场景": scene_key,
            "标签判断总次数": n,
            "不一致次数": mismatch,
            "总体不一致率": safe_div(mismatch, n),
            "误扣次数": fp,
            "误扣率_占全部标签判断": safe_div(fp, n),
            "漏扣次数": fn,
            "漏扣率_占全部标签判断": safe_div(fn, n),
            "平均标签准确率": 1 - safe_div(mismatch, n) if n else np.nan,
        })

    scene_summary_df = pd.DataFrame(scene_rows)

    # 排序：优先显示不一致率高的问题
    label_summary_df = label_summary_df.sort_values(
        by=["场景", "不一致率", "不一致次数"],
        ascending=[True, False, False]
    ).reset_index(drop=True)

    detail_df = detail_df.sort_values(
        by=["场景", "扣分项", "house_id"]
    ).reset_index(drop=True)

    return detail_df, label_summary_df, scene_summary_df, skip_df


def format_excel(output_path: Path, sheets: dict):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if df is None or len(df) == 0:
                df = pd.DataFrame({"说明": ["无数据"]})
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 简单美化
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(output_path)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col_cells:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                width = min(max(max_len + 2, 10), 42)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # 百分比列格式
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    header = ws.cell(row=1, column=cell.column).value
                    if header and ("率" in str(header) or str(header) in ["Precision_扣分精确率", "Recall_扣分召回率", "F1"]):
                        cell.number_format = "0.00"
                    elif header and ("概率" in str(header) or str(header) == "阈值"):
                        cell.number_format = "0.000"

        wb.save(output_path)
    except Exception as e:
        print("Excel 美化失败，不影响结果文件:", e)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=str, default="data/raw", help="农户文件夹根目录")
    parser.add_argument("--output", type=str, default="outputs/label_mismatch_report.xlsx", help="Excel 输出路径")
    args = parser.parse_args()

    root_dir = Path(args.root)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    detail_df, label_summary_df, scene_summary_df, skip_df = evaluate(root_dir)

    format_excel(output_path, {
        "扣分项不一致率": label_summary_df,
        "场景整体不一致率": scene_summary_df,
        "每户每标签明细": detail_df,
        "跳过样本": skip_df,
    })

    print("\n评价完成。")
    print("Excel 已保存:", output_path)
    print("\n重点查看：")
    print("1. 扣分项不一致率：每个场景每个问题的 AI 与人工不符合概率")
    print("2. 场景整体不一致率：每个场景所有问题的整体不一致情况")
    print("3. 每户每标签明细：可追溯到具体农户、具体图片、具体标签")


if __name__ == "__main__":
    main()
